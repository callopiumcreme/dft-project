#!/usr/bin/env python3
"""Generate Zuniga-bascula style ticket xlsx files for DFT 2025 daily inputs.
Pixel-faithful to original template (merges, alignments, row heights, fonts, page setup).

One ticket per delivery (= one daily_inputs row). Driver (COND), plate (TECERO)
and departure date/time are pulled from the SAME deterministic source used by the
eRSV renderer — ``app.services.ersv_pool.build_pool_fields`` — keyed by
(entry_date, position_in_day, total_in_day). This guarantees the conductor on the
bascula ticket matches the conductor named on that delivery's eRSV exactly."""
import os, sys, random, subprocess, datetime as dt
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment

# Reuse the eRSV pool verbatim (no DB deps) so ticket COND == eRSV Nombre conductor.
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))
from app.services.ersv_pool import build_pool_fields  # noqa: E402

random.seed(42)  # only governs bascula-operator (WEIGHER) + peso/tare jitter

OUT_DIR = "/tmp/dft_tickets_out"
os.makedirs(OUT_DIR, exist_ok=True)

FONT_BOLD = Font(name="Courier New", size=12, bold=True)
AL_C = Alignment(horizontal="center")
AL_L = Alignment(horizontal="left")

# Per-ticket template: list of dicts. Indexed by relative row (1..44).
# Each row: a, b (values; can be placeholder strings), merge (bool A:B), align_a, align_b, height
# Heights from original: 5-10=16.2, 11-12=15.6, 13-44=16.2. Rows 1-4 default (merged).
def template_rows():
    h_def = None
    h_main = 16.2
    h_gap = 15.6
    return [
        # row 1-4: ZUNIGA merged across 4 rows
        {"a": "ZUNIGA MARTINEZ S.A.S", "b": None, "align_a": AL_C, "height": h_def, "_merge_block": "header"},
        {"a": None, "b": None, "height": h_def, "_merge_block": "header"},
        {"a": None, "b": None, "height": h_def, "_merge_block": "header"},
        {"a": None, "b": None, "height": h_def, "_merge_block": "header"},
        # 5-10: address header lines, each merged A:B center
        {"a": "M 1.5 VIA CAVASA DE CANDELAR:", "b": None, "merge": True, "align_a": AL_C, "height": h_main},
        {"a": 4359423, "b": None, "merge": True, "align_a": AL_C, "height": h_main},
        {"a": "CANDELARIA ", "b": None, "merge": True, "align_a": AL_C, "height": h_main},
        {"a": "COLOMBIA ", "b": None, "merge": True, "align_a": AL_C, "height": h_main},
        {"a": "REGISTRO DE SERVICIO DE ", "b": None, "merge": True, "align_a": AL_C, "height": h_main},
        {"a": "BASCULA ", "b": None, "merge": True, "align_a": AL_C, "height": h_main},
        # 11-12: gap rows
        {"a": None, "b": None, "height": h_gap},
        {"a": None, "b": None, "height": h_gap},
        # 13: NUM TIQUETE
        {"a": "  NUM. TIQUETE:", "b": "{NUM}", "align_a": AL_C, "align_b": AL_C, "height": h_main},
        {"a": None, "b": None, "height": h_main},
        # 15: ENTRADA / PESO ENT header
        {"a": "  ENTRADA:", "b": "PESO ENT", "align_a": AL_C, "align_b": AL_C, "height": h_main},
        # 16: peso ent value (col B only)
        {"a": None, "b": "{PESO_ENT}", "align_b": AL_C, "height": h_main},
        # 17: FECHA ENT / HORA ENT labels
        {"a": "FECHA ENT ", "b": "HORA ENT", "align_a": AL_C, "align_b": AL_C, "height": h_main},
        # 18: date / time
        {"a": "{FECHA}", "b": "{HORA_ENT}", "align_a": AL_C, "align_b": AL_C, "height": h_main, "_is_time_b": True},
        {"a": None, "b": None, "height": h_main},
        # 20: SALIDA / PESO SAL
        {"a": "   SALIDA:", "b": "PESO SAL", "align_a": AL_C, "align_b": AL_C, "height": h_main},
        {"a": None, "b": "{PESO_SAL}", "align_b": AL_C, "height": h_main},
        {"a": "FECHA SAL", "b": "HORA SAL", "align_a": AL_C, "align_b": AL_C, "height": h_main},
        {"a": "{FECHA}", "b": "{HORA_SAL}", "align_a": AL_C, "align_b": AL_C, "height": h_main, "_is_time_b": True},
        # 24: PESO NETO
        {"a": "PESO NETO:", "b": "{PESO_NETO}", "align_a": AL_C, "align_b": AL_C, "height": h_main},
        {"a": None, "b": None, "height": h_main},
        {"a": None, "b": None, "height": h_main},
        # 27: COND
        {"a": "          COND: {COND}", "b": None, "merge": True, "align_a": AL_L, "height": h_main},
        {"a": None, "b": None, "height": h_main},
        {"a": None, "b": None, "height": h_main},
        # 30: TECERO
        {"a": "        TECERO: {PLACA}", "b": None, "merge": True, "align_a": AL_L, "height": h_main},
        # 31: PROD
        {"a": "          PROD: {PROD}", "b": None, "merge": True, "align_a": AL_L, "height": h_main},
        # 32: TARIFA
        {"a": "        TARIFA: TARIFA UNICA ", "b": None, "merge": True, "align_a": AL_L, "height": h_main},
        # 33-40: blank
        *[{"a": None, "b": None, "height": h_main} for _ in range(8)],
        # 41: separator
        {"a": "      ____________________________________", "b": None, "align_a": AL_L, "height": h_main},
        # 42: PESADO POR
        {"a": "PESADO POR ", "b": None, "merge": True, "align_a": AL_C, "height": h_main},
        # 43: weigher name
        {"a": "{WEIGHER}", "b": None, "merge": True, "align_a": AL_C, "height": h_main},
        # 44: blank merged
        {"a": None, "b": None, "merge": True, "height": h_main},
    ]


# ----- Pools -----
# Driver + plate now come from ersv_pool (== eRSV). Only the bascula operator
# (PESADO POR) has no eRSV counterpart, so it stays a local pool.
WEIGHERS = ["Carlos Hernandez", "Manuel Solorzano", "Ramiro Pena", "Liliana Fajardo"]

# ----- Fetch data -----
# One row = one delivery = one eRSV. entry_time is the real bascula arrival.
sql = """COPY (
  SELECT id, entry_date::text, entry_time::text, total_input_kg::float,
         car_kg::float, truck_kg::float, special_kg::float,
         COALESCE(ersv_number, '') AS ersv_number,
         (SELECT code FROM suppliers s WHERE s.id = di.supplier_id) AS supplier_code
  FROM daily_inputs di
  WHERE deleted_at IS NULL
    AND entry_date BETWEEN '2025-01-01' AND '2025-09-30'
  ORDER BY entry_date, id
) TO STDOUT WITH (FORMAT csv, HEADER true)"""
DB_CONTAINER = os.environ.get("DB_CONTAINER", "dft-project-db-1")
out = subprocess.check_output(
    ["docker", "exec", DB_CONTAINER, "psql", "-U", "dft", "-d", "dft", "-c", sql],
    text=True,
)
import csv, io
rows = list(csv.DictReader(io.StringIO(out)))
print(f"loaded {len(rows)} daily_inputs rows")


def _parse_time(s):
    """psql time text 'HH:MM:SS' -> datetime.time, or None."""
    if not s:
        return None
    parts = s.split(":")
    return dt.time(int(parts[0]), int(parts[1]))


def _prod_for(car, truck, special):
    """eRSV declares ELT (neumaticos) for tyre loads. 'special' = non-tyre lot."""
    if special > 0 and car == 0 and truck == 0:
        return "MIXTO PLASTICO"
    return "LLANTAS"


# One ticket per daily_inputs row. position_in_day / total_in_day computed
# exactly as ersv_renderer.py (rank by id within active rows of that date) so
# build_pool_fields returns the SAME driver/plate as that row's eRSV.
tickets_per_day = defaultdict(list)
rows_by_day = defaultdict(list)
for r in rows:
    rows_by_day[dt.date.fromisoformat(r["entry_date"])].append(r)

ticket_num = 24500
for d in sorted(rows_by_day.keys()):
    day_rows = sorted(rows_by_day[d], key=lambda r: int(r["id"]))
    total_in_day = len(day_rows)
    for pos, r in enumerate(day_rows, start=1):
        car, truck, special = float(r["car_kg"]), float(r["truck_kg"]), float(r["special_kg"])
        net = float(r["total_input_kg"])
        pool = build_pool_fields(
            r["ersv_number"],
            d,
            daily_input_id=int(r["id"]),
            position_in_day=pos,
            total_in_day=total_in_day,
            supplier_code=r["supplier_code"],
        )
        rrng = random.Random(int(r["id"]))  # per-row reproducible jitter
        hora_ent = _parse_time(r["entry_time"]) or dt.time(
            rrng.randint(5, 11), rrng.randint(0, 59)
        )
        sal_min = (hora_ent.hour * 60 + hora_ent.minute) + rrng.randint(60, 120)
        hora_sal = dt.time((sal_min // 60) % 24, sal_min % 60)
        tare = rrng.randint(14000, 18000)
        tickets_per_day[d].append({
            "num": ticket_num,
            "ersv": r["ersv_number"],
            "net_kg": net,
            "prod": _prod_for(car, truck, special),
            "driver": pool["driver_name"],        # == eRSV Nombre conductor
            "plate": pool["vehicle_plate"],        # == eRSV Placa vehiculo
            "weigher": rrng.choice(WEIGHERS),
            "hora_ent": hora_ent,                  # real bascula arrival (entry_time)
            "hora_sal": hora_sal,
            "peso_sal_kg": tare,
            "peso_ent_kg": tare + int(net),
            "peso_neto_kg": int(net),
        })
        ticket_num += 1
    tickets_per_day[d].sort(key=lambda x: x["hora_ent"])

total_tickets = sum(len(v) for v in tickets_per_day.values())
print(f"days: {len(tickets_per_day)}, total tickets: {total_tickets}")


def fmt_kg(n, suffix):
    """Format like '32.999 Kg.' — thousands sep is '.' (Spanish/Colombian)."""
    return f"{n:,}{suffix}".replace(",", ".")


def substitute(template_value, t, fecha_str):
    if not isinstance(template_value, str):
        return template_value
    return (
        template_value
        .replace("{NUM}", str(t["num"]))
        .replace("{PESO_ENT}", fmt_kg(t["peso_ent_kg"], " Kg."))
        .replace("{PESO_SAL}", fmt_kg(t["peso_sal_kg"], " kg"))
        .replace("{PESO_NETO}", fmt_kg(t["peso_neto_kg"], " Kg"))
        .replace("{FECHA}", fecha_str)
        .replace("{COND}", t["driver"])
        .replace("{PLACA}", t["plate"])
        .replace("{PROD}", t["prod"])
        .replace("{WEIGHER}", t["weigher"])
    )


def write_ticket(ws, start_row, t, fecha_str):
    """Write one ticket starting at start_row. Returns end_row+1 (start of next ticket)."""
    template = template_rows()
    for offset, spec in enumerate(template):
        r = start_row + offset
        a_val = substitute(spec.get("a"), t, fecha_str)
        b_val = spec.get("b")
        if isinstance(b_val, str) and "{HORA_ENT}" in b_val:
            b_val = t["hora_ent"]
        elif isinstance(b_val, str) and "{HORA_SAL}" in b_val:
            b_val = t["hora_sal"]
        else:
            b_val = substitute(b_val, t, fecha_str)
        # write values
        ca = ws.cell(row=r, column=1, value=a_val)
        cb = ws.cell(row=r, column=2, value=b_val)
        if a_val is not None:
            ca.font = FONT_BOLD
        if b_val is not None:
            cb.font = FONT_BOLD
            if isinstance(b_val, dt.time):
                cb.number_format = "hh:mm AM/PM"
        # alignment
        if spec.get("align_a"):
            ca.alignment = spec["align_a"]
        if spec.get("align_b"):
            cb.alignment = spec["align_b"]
        # row height
        if spec.get("height"):
            ws.row_dimensions[r].height = spec["height"]
    # apply merges
    # header block: rows 1-4 merged A:B
    ws.merge_cells(start_row=start_row, end_row=start_row + 3, start_column=1, end_column=2)
    # per-row merges where spec.merge=True
    for offset, spec in enumerate(template):
        if spec.get("merge"):
            r = start_row + offset
            ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=2)
    return start_row + len(template)


# Group by month + write
months = defaultdict(list)
for d in sorted(tickets_per_day.keys()):
    months[(d.year, d.month)].append(d)

for (y, m), days in sorted(months.items()):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for d in days:
        sn = d.strftime("%d-%m-%y")
        ws = wb.create_sheet(sn)
        # widths
        ws.column_dimensions["A"].width = 28.44
        ws.column_dimensions["B"].width = 51.88
        ws.column_dimensions["C"].width = 11.55
        # page setup
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = 275
        ws.page_margins.left = 0.11
        ws.page_margins.right = 0.06
        ws.page_margins.top = 0.2
        ws.page_margins.bottom = 0.12
        ws.page_margins.header = 0.0
        ws.page_margins.footer = 0.0
        # tickets stacked vertically
        next_row = 1
        fecha_str = d.strftime("%d-%b-%Y")
        for t in tickets_per_day[d]:
            next_row = write_ticket(ws, next_row, t, fecha_str)
    fname = f"TICKETS_DFT_2025_{m:02d}.xlsx"
    fpath = os.path.join(OUT_DIR, fname)
    wb.save(fpath)
    n = sum(len(tickets_per_day[d]) for d in days)
    print(f"wrote {fpath} ({len(days)} sheets, {n} tickets)")

print(f"\nDONE. ticket range: 24500 .. {ticket_num - 1}")

