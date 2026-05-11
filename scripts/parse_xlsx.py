"""Parser xlsx Girardot — convert mixed-layout sheet to daily_inputs[] + daily_production[].

Layout per sheet (22 cols A-V):
  R16          : 'LOADING SUMMARY' marker — content starts after.
  Date row     : col A string 'DD MONTH YYYY'. Optional STOCK marker in cols I/J
                 (ignored — stock balance recomputed from inputs - production).
                 May also carry aggregate prod fields (K+L+M+O populated).
  Header row   : col A = 'TIME'.
  Transaction  : col A = datetime.time, col B = supplier name.
                 May co-carry per-day byproduct breakdown in cols O-S (first batch
                 row of the day typically holds the full daily breakdown).
  TOTAL row    : col E = 'TOTAL' — skipped.
  Aggregate row: col A empty, K+L+M+O populated. On aggregate, O = SUM of
                 byproducts (carbon + metal + H2O + gas + losses), NOT carbon alone.
                 Used to set kg_to_production / eu_prod_kg / plus_prod_kg.
  Detail row   : O+P+Q+R+S all populated = per-day byproduct breakdown.
                 Co-occurs with transaction (A=time) or standalone (A=empty).
                 Used to set carbon/metal/h2o/gas/losses individually.
  Stock recov. : col A empty, K populated alone (no L/M/O detail). Internal
                 accounting movement — skipped (stock recomputed downstream).

Sign conventions (xlsx):
  kg_to_production and byproduct kg are negative in xlsx (loss accounting).
  Stored as positive (abs) — DB schema treats as positive quantities.

When same date has multiple aggregate rows (e.g. r49 date-row + r62 A=None for
day 7 in Girardot Jan-2025), LAST-WINS: later row supersedes earlier.
"""
from __future__ import annotations

import re
import sys
from dataclasses import asdict, dataclass
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook

MONTH_MAP = {
    "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4,
    "MAY": 5, "JUNE": 6, "JULY": 7, "AUGUST": 8,
    "SEPTEMBER": 9, "OCTOBER": 10, "NOVEMBER": 11, "DECEMBER": 12,
}
DATE_RE = re.compile(r"^\s*(\d{1,2})\s+([A-Z]+)\s+(\d{4})\s*$")
SUPPLIER_ALIAS = {
    "≤5 TON": "LE5TON",
    "<=5 TON": "LE5TON",
    "<5 TON": "LE5TON",
}


@dataclass
class DailyInput:
    entry_date: date
    entry_time: time | None
    supplier_code: str
    cert_number: str | None
    contract_code: str | None
    ersv_number: str | None
    car_kg: float
    truck_kg: float
    special_kg: float
    theor_veg_pct: float | None
    manuf_veg_pct: float | None
    c14_analysis: str | None
    source_file: str
    source_row: int


@dataclass
class DailyProduction:
    prod_date: date
    kg_to_production: float | None = None
    eu_prod_kg: float | None = None
    plus_prod_kg: float | None = None
    carbon_black_kg: float | None = None
    metal_scrap_kg: float | None = None
    h2o_kg: float | None = None
    gas_syngas_kg: float | None = None
    losses_kg: float | None = None
    output_eu_kg: float | None = None
    contract_ref: str | None = None
    pos_number: str | None = None
    source_file: str = ""
    source_row: int = 0


def _parse_date(s: str) -> date | None:
    m = DATE_RE.match(s.upper())
    if not m:
        return None
    day, month_name, year = m.groups()
    month = MONTH_MAP.get(month_name)
    if not month:
        return None
    try:
        return date(int(year), month, int(day))
    except ValueError:
        return None


def _norm_supplier(name: str) -> str:
    name = name.strip()
    return SUPPLIER_ALIAS.get(name.upper(), name)


def _f(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _abs_or_none(v) -> float | None:
    f = _f(v)
    return abs(f) if f is not None else None


def _as_time(v) -> time | None:
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    return None


def _is_aggregate_prod(row: list) -> bool:
    """K + L + M + O all populated → daily aggregate row.
    On aggregate, O is total byproducts (not carbon alone)."""
    return all(_f(row[i]) is not None for i in (10, 11, 12, 14))


def _is_byproduct_detail(row: list) -> bool:
    """O + P + Q + R + S all populated → per-day byproduct breakdown."""
    return all(_f(row[i]) is not None for i in (14, 15, 16, 17, 18))


def _resolve_aggregate_date(
    d: date, prev_date: date | None, row: list, productions: dict
) -> date:
    """When a date-row carries an embedded aggregate, decide whether it belongs
    to d (the date label) or to prev_date. Heuristic: if prev_date has orphan
    byproducts (no kg_to_production yet) whose sum matches this aggregate's
    col O within 2 kg, the aggregate is actually for prev_date.

    Background: in Girardot xlsx, day-6 byproducts (r42) sum to 43,897 which
    matches col O of r49 (labeled '07 JAN 2025' but K+L+M+O actually belong to
    day 6). Days with their own separate aggregate row do not trigger this."""
    if prev_date is None or prev_date not in productions:
        return d
    prev = productions[prev_date]
    if prev.kg_to_production is not None:
        return d
    by_sum = sum(
        v for v in (
            prev.carbon_black_kg, prev.metal_scrap_kg, prev.h2o_kg,
            prev.gas_syngas_kg, prev.losses_kg,
        ) if v is not None
    )
    if by_sum <= 0:
        return d
    agg_o = abs(float(row[14]))
    return prev_date if abs(by_sum - agg_o) <= 2.0 else d


def _find_monthly_total_row(ws) -> int:
    """Locate the 'TOTAL PROCESSED' header row that marks end-of-data per sheet.
    The row immediately above it is the month-aggregate row (K = sum of all
    daily K) and must not be parsed as a daily aggregate. Returns the row
    number of the marker (exclusive upper bound) or ws.max_row + 1."""
    for r in range(17, ws.max_row + 1):
        for c in range(1, 23):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "TOTAL PROCESSED" in v.upper():
                # back up past the preceding aggregate-total row (typically 1 row up)
                # and any TOTAL/blank rows above it
                cutoff = r - 1
                while cutoff > 17:
                    above = [ws.cell(cutoff, c).value for c in range(1, 23)]
                    if _is_aggregate_prod(above) and above[0] is None:
                        cutoff -= 1
                        continue
                    break
                return cutoff + 1
    return ws.max_row + 1


def parse_workbook(path: Path) -> tuple[list[DailyInput], list[DailyProduction]]:
    wb = load_workbook(path, data_only=True)
    inputs: list[DailyInput] = []
    productions: dict[date, DailyProduction] = {}
    anomalies: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        src = f"{path.name}#{sheet_name}"
        current_date: date | None = None
        end_row = _find_monthly_total_row(ws)
        for r in range(1, end_row):
            row = [ws.cell(r, c).value for c in range(1, 23)]
            if all(v is None or v == "" for v in row):
                continue

            a = row[0]
            e = row[4]

            if isinstance(a, str):
                d = _parse_date(a)
                if d:
                    prev_date = current_date
                    current_date = d
                    if _is_aggregate_prod(row):
                        target = _resolve_aggregate_date(d, prev_date, row, productions)
                        prod = productions.setdefault(
                            target, DailyProduction(prod_date=target, source_file=src, source_row=r)
                        )
                        _merge_aggregate(prod, row, r)
                    continue
                if a.strip().upper() == "TIME":
                    continue

            if isinstance(e, str) and e.strip().upper() == "TOTAL":
                if current_date is not None and _is_aggregate_prod(row):
                    prod = productions.setdefault(
                        current_date,
                        DailyProduction(prod_date=current_date, source_file=src, source_row=r),
                    )
                    _merge_aggregate(prod, row, r)
                continue

            t = _as_time(a)
            if t is not None:
                if current_date is None:
                    continue
                if row[1]:
                    inputs.append(_make_input(current_date, t, row, src, r))
                if _is_byproduct_detail(row):
                    prod = productions.setdefault(
                        current_date,
                        DailyProduction(prod_date=current_date, source_file=src, source_row=r),
                    )
                    _merge_byproducts(prod, row, r)
                continue

            if row[1] and current_date is not None:
                sup_raw = str(row[1]).strip().upper()
                if sup_raw in SUPPLIER_ALIAS:
                    inputs.append(_make_input(current_date, None, row, src, r))
                    if _is_byproduct_detail(row):
                        prod = productions.setdefault(
                            current_date,
                            DailyProduction(prod_date=current_date, source_file=src, source_row=r),
                        )
                        _merge_byproducts(prod, row, r)
                    continue

            if (a is None or a == "") and current_date is not None:
                if _is_aggregate_prod(row):
                    prod = productions.setdefault(
                        current_date,
                        DailyProduction(prod_date=current_date, source_file=src, source_row=r),
                    )
                    _merge_aggregate(prod, row, r)
                    continue
                if _is_byproduct_detail(row):
                    prod = productions.setdefault(
                        current_date,
                        DailyProduction(prod_date=current_date, source_file=src, source_row=r),
                    )
                    _merge_byproducts(prod, row, r)
                    continue
                if _f(row[10]) is not None:
                    anomalies.append(
                        f"{src} r{r} date={current_date} partial-prod-or-stock"
                        f" K={row[10]} L={row[11]} M={row[12]} O={row[14]} P={row[15]}"
                    )
                    continue

    if anomalies:
        print(
            f"\n!! Parser anomalies (skipped, {len(anomalies)} rows):",
            file=sys.stderr,
        )
        for line in anomalies:
            print(f"   {line}", file=sys.stderr)

    return inputs, list(productions.values())


def _make_input(d: date, t: time | None, row: list, src: str, r: int) -> DailyInput:
    return DailyInput(
        entry_date=d,
        entry_time=t,
        supplier_code=_norm_supplier(str(row[1])),
        cert_number=str(row[2]).strip() if row[2] else None,
        contract_code=str(row[3]).strip() if row[3] else None,
        ersv_number=str(row[4]).strip() if row[4] else None,
        car_kg=_f(row[5]) or 0.0,
        truck_kg=_f(row[6]) or 0.0,
        special_kg=_f(row[7]) or 0.0,
        theor_veg_pct=_f(row[8]),
        manuf_veg_pct=_f(row[9]),
        c14_analysis=str(row[13]).strip() if row[13] else None,
        source_file=src,
        source_row=r,
    )


def _merge_aggregate(prod: DailyProduction, row: list, r: int) -> None:
    """Set K (input), L (EU), M (PLUS) from aggregate row. Last-wins on duplicate
    aggregates per date. Skips O (total byproducts on aggregate)."""
    new_kg = _abs_or_none(row[10])
    if new_kg is not None:
        prod.kg_to_production = new_kg
    new_eu = _abs_or_none(row[11])
    if new_eu is not None:
        prod.eu_prod_kg = new_eu
    new_plus = _abs_or_none(row[12])
    if new_plus is not None:
        prod.plus_prod_kg = new_plus
    prod.source_row = r


def _merge_byproducts(prod: DailyProduction, row: list, r: int) -> None:
    """Set O/P/Q/R/S as carbon/metal/h2o/gas/losses respectively. Also T (output_eu),
    U (contract_ref), V (pos_number) when present. Last-wins on duplicates."""
    o = _abs_or_none(row[14])
    if o is not None:
        prod.carbon_black_kg = o
    p = _abs_or_none(row[15])
    if p is not None:
        prod.metal_scrap_kg = p
    q = _abs_or_none(row[16])
    if q is not None:
        prod.h2o_kg = q
    gas = _abs_or_none(row[17])
    if gas is not None:
        prod.gas_syngas_kg = gas
    loss = _abs_or_none(row[18])
    if loss is not None:
        prod.losses_kg = loss
    out_eu = _abs_or_none(row[19])
    if out_eu is not None and prod.output_eu_kg is None:
        prod.output_eu_kg = out_eu
    if row[20] and not prod.contract_ref:
        prod.contract_ref = str(row[20]).strip()
    if row[21] and not prod.pos_number:
        prod.pos_number = str(row[21]).strip()


if __name__ == "__main__":
    p = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/tmp/girardot_enero_2025.xlsx")
    inputs, prods = parse_workbook(p)
    print(f"inputs:      {len(inputs)}")
    print(f"productions: {len(prods)}")
    if inputs:
        dates = sorted({i.entry_date for i in inputs})
        print(f"date range:   {dates[0]} -> {dates[-1]}  ({len(dates)} unique days)")
        suppliers = sorted({i.supplier_code for i in inputs})
        print(f"suppliers:    {suppliers}")
    if prods:
        n_with_eu = sum(1 for p in prods if p.eu_prod_kg)
        n_with_carbon = sum(1 for p in prods if p.carbon_black_kg)
        n_with_metal = sum(1 for p in prods if p.metal_scrap_kg)
        print(f"prods w/ eu: {n_with_eu}/{len(prods)}")
        print(f"prods w/ carbon: {n_with_carbon}/{len(prods)}")
        print(f"prods w/ metal_scrap: {n_with_metal}/{len(prods)}")
        for p in sorted(prods, key=lambda x: x.prod_date)[:5]:
            print(f"  {asdict(p)}")
