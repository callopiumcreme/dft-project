"""Render Annex D — Closing Stock Carry-over Declaration for RTFO-310125 bundle.

Story: E1-S1.8 / DFTEN-102. Produces
``deliverables/RTFO-310125/01_annex_a_mass_balance/07_stock_carryover_explanation.pdf``
documenting the 339 865 kg of ELT-derived intermediate physically retained at
the OisteBio Girardot facility at the close of January 2025.

This is the PRODUCTION renderer (sibling to the earlier
``render_stock_carryover_sample.py`` which used hand-picked synthetic per-row
values). The production variant differs on three points:

1. The headline 339 865 kg figure is verified by direct query against the
   live DFT production database — it is computed as
   ``SUM(daily_inputs.total_input_kg) - SUM(daily_production.kg_to_production)``
   for the January 2025 reporting period (filtered on ``deleted_at IS NULL``).
   The mirror figure for February 2025 confirms perfect symmetric carry-over
   (Σ Jan + Σ Feb closure deltas = 0).

2. The K-only stock-recovery rows enumerated on Page 3 reflect REAL values
   read from the operator source workbook
   ``Girardot producciòn Enero 2025_KNEG.xlsx``, sheet ``JANUARY2025``.
   There are 21 such rows (not 17 as in the earlier sample); their sum is
   −774 763 kg, which is the period-level stock withdrawal — NOT the
   period-level closing-stock residual. The two figures are deliberately
   distinct quantities and the template explains the relationship between
   them (§4 row-detail note).

3. Audit-trail anchoring cites the ``audit_log`` id range and timestamp
   window read live from the DFT database (no fabricated identifiers).

Usage:
    python3 scripts/render_stock_carryover.py [out_pdf]

Default output:
    deliverables/RTFO-310125/01_annex_a_mass_balance/07_stock_carryover_explanation.pdf

Constraints:
- DO NOT commit. DO NOT push. (Per Day 6 ops protocol.)
- DO NOT alter the database — read-only operations only.
- Deterministic PDF: same DB state + same xlsx + same context = same SHA-256.
"""

from __future__ import annotations

import datetime as _dt
import hashlib
import re
import subprocess
import sys
from pathlib import Path
from typing import TYPE_CHECKING, Any, Final

from openpyxl import load_workbook

if TYPE_CHECKING:
    from jinja2 import Environment

    from app.services.pdf_renderer import RenderResult  # type: ignore[import-not-found]

_REPO_ROOT: Final[Path] = Path(__file__).resolve().parents[1]
_BACKEND_DIR: Final[Path] = _REPO_ROOT / "backend"
if str(_BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(_BACKEND_DIR))

from app.services import pdf_renderer  # type: ignore[import-not-found]  # noqa: E402

_ORIGINAL_BUILD_ENV: Final = pdf_renderer._build_env

TEMPLATE_NAME: Final[str] = "stock_carryover_v1.html"
DEFAULT_OUTPUT: Final[Path] = (
    _REPO_ROOT
    / "deliverables"
    / "RTFO-310825"
    / "06_annex_d_stock_carryover"
    / "07_stock_carryover_jan_feb_2025.pdf"
)

SOURCE_XLSX: Final[Path] = Path("/tmp/jan2025.xlsx")  # operator workbook, copied locally
JAN_SHEET: Final[str] = "JANUARY2025"

# Container + database connection — read-only.
DB_CONTAINER: Final[str] = "dft-project_db_1"
DB_USER: Final[str] = "dft"
DB_NAME: Final[str] = "dft"

# Headline figure expected from live DB query — pinned as constant for
# defence-in-depth assertion only. The renderer queries the DB at run time
# and refuses to proceed if the live value drifts.
EXPECTED_CLOSING_STOCK_KG: Final[float] = 339_865.0

# Parser fix anchor — see project memory project_xlsx_aggregate_detail_fix.
PARSER_FIX_COMMIT: Final[str] = "99e4edf"
PARSER_FIX_DATE_LABEL: Final[str] = "2026-05-11"
INGEST_DATE_LABEL: Final[str] = "2026-05-11"


# ---------------------------------------------------------------------------
# Jinja numeric filters — match render_stock_carryover_sample.py conventions.
# ---------------------------------------------------------------------------
def _format_thousands(value: float, decimals: int = 2) -> str:
    rounded = round(float(value), decimals)
    formatted = f"{rounded:,.{decimals}f}"
    return formatted.replace(",", " ").replace(".", ",")


def fmt_num(value: float | int | None) -> str:
    if value is None:
        return "—"
    return _format_thousands(float(value), decimals=3)


def fmt_kg(value: float | int | None) -> str:
    if value is None:
        return "—"
    return f"{_format_thousands(float(value), decimals=3)} kg"


def fmt_pct(value: float | int | None) -> str:
    if value is None:
        return "—"
    return f"{_format_thousands(float(value), decimals=2)} %"


# ---------------------------------------------------------------------------
# DB queries — read-only via docker exec psql.
# ---------------------------------------------------------------------------
def _psql_scalar(query: str) -> str:
    """Run ``psql -t -A`` and return the raw single-line result."""
    cmd = ["docker", "exec", DB_CONTAINER, "psql", "-U", DB_USER, "-d", DB_NAME, "-t", "-A", "-c", query]
    out = subprocess.run(cmd, check=True, capture_output=True, text=True)
    return out.stdout.strip()


def fetch_closure_math() -> dict[str, float]:
    """Return Σ inputs / Σ kg_to_production for Jan and Feb 2025 from live DB."""
    jan_input = float(_psql_scalar(
        "SELECT COALESCE(SUM(total_input_kg), 0) FROM daily_inputs "
        "WHERE entry_date BETWEEN '2025-01-01' AND '2025-01-31' AND deleted_at IS NULL;"
    ))
    jan_kp = float(_psql_scalar(
        "SELECT COALESCE(SUM(kg_to_production), 0) FROM daily_production "
        "WHERE prod_date BETWEEN '2025-01-01' AND '2025-01-31' AND deleted_at IS NULL;"
    ))
    feb_input = float(_psql_scalar(
        "SELECT COALESCE(SUM(total_input_kg), 0) FROM daily_inputs "
        "WHERE entry_date BETWEEN '2025-02-01' AND '2025-02-28' AND deleted_at IS NULL;"
    ))
    feb_kp = float(_psql_scalar(
        "SELECT COALESCE(SUM(kg_to_production), 0) FROM daily_production "
        "WHERE prod_date BETWEEN '2025-02-01' AND '2025-02-28' AND deleted_at IS NULL;"
    ))
    return {
        "jan_input_kg": jan_input,
        "jan_kg_to_production": jan_kp,
        "jan_delta_kg": jan_input - jan_kp,
        "feb_input_kg": feb_input,
        "feb_kg_to_production": feb_kp,
        "feb_delta_kg": feb_input - feb_kp,
        "combined_input_kg": jan_input + feb_input,
        "combined_kg_to_production": jan_kp + feb_kp,
        "combined_delta_kg": (jan_input - jan_kp) + (feb_input - feb_kp),
    }


def fetch_audit_log_anchor() -> dict[str, Any]:
    """Return audit_log range/window for tables instrumented during re-ingest."""
    raw = _psql_scalar(
        "SELECT COALESCE(MIN(id), 0), COALESCE(MAX(id), 0), "
        "COALESCE(MIN(changed_at)::text, ''), COALESCE(MAX(changed_at)::text, ''), "
        "COUNT(*) FROM audit_log "
        "WHERE table_name IN ('daily_inputs','daily_production');"
    )
    parts = raw.split("|")
    if len(parts) != 5:
        msg = f"unexpected audit_log row shape: {raw!r}"
        raise RuntimeError(msg)
    id_min, id_max, ts_min, ts_max, n = parts
    return {
        "audit_log_id_min": id_min,
        "audit_log_id_max": id_max,
        "audit_log_window_min": ts_min,
        "audit_log_window_max": ts_max,
        "audit_log_entries_count": int(n),
    }


# ---------------------------------------------------------------------------
# Source xlsx — pull the K-only stock-recovery rows.
# ---------------------------------------------------------------------------
_DATE_RE: Final = re.compile(r"^(\d{1,2})\s+([A-Z]+)\s+(\d{4})", re.IGNORECASE)


def _attribute_day_labels(ws: Any) -> dict[int, str]:
    """Return row → last-seen day label scanning cols A–D for date strings/objects."""
    last: str | None = None
    out: dict[int, str | None] = {}
    for r in range(1, ws.max_row + 1):
        for col in range(1, 5):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, (_dt.date, _dt.datetime)):
                last = v.strftime("%Y-%m-%d")
                break
            if isinstance(v, str) and _DATE_RE.match(v.strip()):
                last = v.strip()
                break
        out[r] = last
    return {r: (label or "(no date)") for r, label in out.items()}


def fetch_kcol_only_rows() -> list[dict[str, Any]]:
    """Return the list of K-only rows from the JANUARY2025 sheet."""
    if not SOURCE_XLSX.exists():
        msg = (
            f"Source xlsx not found at {SOURCE_XLSX}. Copy the operator workbook "
            "to that path before re-rendering Annex D."
        )
        raise FileNotFoundError(msg)
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb[JAN_SHEET]
    day_labels = _attribute_day_labels(ws)
    rows: list[dict[str, Any]] = []
    subtotal = 0.0
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        k = ws.cell(row=r, column=11).value
        l_val = ws.cell(row=r, column=12).value
        m_val = ws.cell(row=r, column=13).value
        o_val = ws.cell(row=r, column=15).value
        if a is None and isinstance(k, (int, float)) and l_val is None and m_val is None and o_val is None:
            kg = float(k)
            subtotal += kg
            rows.append({
                "xlsx_row": r,
                "date_label": day_labels.get(r) or "(no date)",
                "kg": kg,
                "subtotal": subtotal,
            })
    return rows


# ---------------------------------------------------------------------------
# Context assembly + render.
# ---------------------------------------------------------------------------
def build_context() -> dict[str, Any]:
    closure = fetch_closure_math()
    audit = fetch_audit_log_anchor()
    kcol_rows = fetch_kcol_only_rows()

    closing_stock_kg = closure["jan_delta_kg"]
    if round(closing_stock_kg, 3) != round(EXPECTED_CLOSING_STOCK_KG, 3):
        msg = (
            f"closing-stock from live DB ({closing_stock_kg} kg) differs from "
            f"expected canonical figure ({EXPECTED_CLOSING_STOCK_KG} kg). "
            "Refusing to render Annex D without manual review."
        )
        raise RuntimeError(msg)

    kcol_total = round(sum(r["kg"] for r in kcol_rows), 3)

    return {
        "submission_ref": "RTFO-310125",
        "period": "2025-01",
        "period_label": "January 2025",
        "generated_at": "2026-05-20T00:00:00Z",
        "submitter_company": "OisteBio GmbH",
        "submitter_address": "Oberneuhofstrasse 5, 6340 Baar, Switzerland",
        "plant_name": "Girardot, Colombia",
        "feedstock": "End-of-Life Tyres (ELT) — pyrolysis intermediate",
        "regulator": "UK Department for Transport — RTFO / LCF Delivery Unit",
        "source_workbook": "Girardot producciòn Enero 2025_KNEG.xlsx",
        "legal_name": "OisteBio GmbH",
        "legal_address": "Oberneuhofstrasse 5, 6340 Baar, Switzerland",
        "legal_vat": "CHE-234.625.162",
        "legal_jurisdiction": "Canton of Zug — Switzerland",
        # Closure math (from live DB).
        "closing_stock_kg": closing_stock_kg,
        "jan_input_kg": closure["jan_input_kg"],
        "jan_kg_to_production": closure["jan_kg_to_production"],
        "jan_delta_kg": closure["jan_delta_kg"],
        "feb_input_kg": closure["feb_input_kg"],
        "feb_kg_to_production": closure["feb_kg_to_production"],
        "feb_delta_kg": closure["feb_delta_kg"],
        "combined_input_kg": closure["combined_input_kg"],
        "combined_kg_to_production": closure["combined_kg_to_production"],
        "combined_delta_kg": closure["combined_delta_kg"],
        # K-only row detail.
        "composition_rows": kcol_rows,
        "kcol_only_row_count": len(kcol_rows),
        "kcol_only_total_kg": kcol_total,
        # Audit-trail anchor.
        "parser_fix_commit": PARSER_FIX_COMMIT,
        "parser_fix_date_label": PARSER_FIX_DATE_LABEL,
        "ingest_date_label": INGEST_DATE_LABEL,
        **audit,
        # Hash self-anchor — placeholder; overwritten on a second render pass.
        "annex_d_hash": "0" * 64,
    }


def _patched_env_factory() -> Environment:
    base_env: Environment = _ORIGINAL_BUILD_ENV()
    base_env.filters["fmt_num"] = fmt_num
    base_env.filters["fmt_kg"] = fmt_kg
    base_env.filters["fmt_pct"] = fmt_pct
    return base_env


def render(out_pdf: Path, context: dict[str, Any]) -> RenderResult:
    pdf_renderer._build_env = _patched_env_factory
    try:
        return pdf_renderer.render_to_pdf(
            template_name=TEMPLATE_NAME,
            context=context,
            output_path=out_pdf,
        )
    finally:
        pdf_renderer._build_env = _ORIGINAL_BUILD_ENV


def main(argv: list[str]) -> int:
    out_pdf = Path(argv[1]) if len(argv) > 1 else DEFAULT_OUTPUT
    out_pdf = out_pdf.resolve()

    # Pass 1 — render with placeholder hash so we can compute the real digest.
    context = build_context()
    result1 = render(out_pdf, context)

    # Pass 2 — re-render with the real digest embedded in the footer so the
    # printed short prefix matches the side-car. This is the same
    # self-anchoring pattern used by render_iscc_pos_status.py.
    context["annex_d_hash"] = result1.pdf_sha256
    result2 = render(out_pdf, context)

    size_kb = result2.pdf_size_bytes / 1024
    on_disk_sha = hashlib.sha256(result2.pdf_path.read_bytes()).hexdigest()

    print(f"Rendered:    {result2.pdf_path}")
    print(f"Size:        {size_kb:.1f} KB ({result2.pdf_size_bytes} bytes)")
    print(f"Pages:       {result2.page_count}")
    print(f"Template:    {result2.template_name}")
    print(f"SHA-256:     {result2.pdf_sha256}")
    print(f"On-disk:     {on_disk_sha}")
    print(f"Side-car:    {result2.pdf_sha256_path}")
    print(f"K-only rows: {context['kcol_only_row_count']} (sum {context['kcol_only_total_kg']} kg)")
    print(f"Closing kg:  {context['closing_stock_kg']}")
    print(f"Audit log:   id {context['audit_log_id_min']}-{context['audit_log_id_max']} "
          f"({context['audit_log_entries_count']} entries)")

    if on_disk_sha != result2.pdf_sha256:
        print("ERROR: on-disk SHA-256 differs from renderer-reported digest", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
